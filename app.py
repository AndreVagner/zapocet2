from flask import Flask, render_template, request, redirect, flash, url_for, jsonify, g, session, Response
from datetime import datetime, timedelta
import hashlib
import os
from functools import wraps
from flask_login import LoginManager, UserMixin, login_user, current_user, logout_user
import sqlite3
from collections import defaultdict
import requests
from flask_caching import Cache
from openpyxl import Workbook
import io
from mailjet_rest import Client

app = Flask(__name__)

api_key = '4920777f5d6f51541b32e6e7acdc7b42'
api_secret = 'cbe47b9e577b785ea8b2b41bd72099eb'
mailjet = Client(auth=(api_key, api_secret), version='v3.1')

# Nastavení tajného klíče pro bezpečnostní účely (např. pro session)
app.secret_key = os.environ.get('SECRET_KEY', 'O<giKIn&F$%!]Odr6r/D8v3?T?kcuA')
app.debug = True

app.config['CACHE_TYPE'] = 'simple'
cache = Cache(app)

# Cesta k databázovému souboru
DATABASE = r'.\finalniDB.db'
#DATABASE = '/home/andvag/BcWk/finalniDB.db'
# Funkce pro připojení k databázi
def connect_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE, timeout=30)
    return db

# Před každým requestem se vytvoří nové databázové spojení a uloží do g
@app.before_request
def before_request():
    g.db = connect_db()

# Po každém requestu se databázové spojení zavře
@app.after_request
def after_request(response):
    g.db.close()
    return response

# Ověření, zda je uživatel přihlášen
def login_nutny(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Konfigurace pro Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Mapování měsíců na jejich číselné hodnoty
mesice = {
    "Leden": 1,
    "Únor": 2,
    "Březen": 3,
    "Duben": 4,
    "Květen": 5,
    "Červen": 6,
    "Červenec": 7,
    "Srpen": 8,
    "Září": 9,
    "Říjen": 10,
    "Listopad": 11,
    "Prosinec": 12
}


def kurzy():
    api_key = '8605ae6f977b8e7eaa4f9bff'
    url = f'https://v6.exchangerate-api.com/v6/{api_key}/latest/CZK'
    kurzy = cache.get('rates')
    if not kurzy:
        response = requests.get(url)
        data = response.json()
        kurzy = data['conversion_rates']
        cache.set('rates', kurzy, timeout=86400)
    return kurzy

def prevest_menu(castka, mena, to_czk=False):
    kurzy2 = kurzy()
    if to_czk:
        return castka / kurzy2.get(mena, 1)
    return castka * kurzy2.get(mena, 1)

@app.route('/', methods=['GET', 'POST'])
@login_nutny
def home():

    aktivni_str = "home"

    if request.method == 'POST':
        session['currency'] = request.form.get('currency', 'CZK')

    # Získání aktuálně vybrané měny
    vybrana_mena = session.get('currency', 'CZK')

    cursor = g.db.cursor()
    cursor.execute('SELECT SUM(Castka) FROM Prijmy WHERE UserID = ?', (current_user.id,))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute('SELECT SUM(Castka) FROM Vydaje WHERE UserID = ?', (current_user.id,))
    total_expense = cursor.fetchone()[0] or 0

    # Přepočet měn
    total_income = prevest_menu(total_income, vybrana_mena)
    total_expense = prevest_menu(total_expense, vybrana_mena)

    total_balance = total_income - total_expense

    return render_template('index.html', total_income=total_income, total_expense=total_expense,
                           total_balance=total_balance, selected_currency=vybrana_mena, aktivni_str = aktivni_str)

@app.route('/incomes', methods=['GET', 'POST'])
@login_nutny
def incomes():

    aktivni_str = "incomes"

    vybrana_mena = session.get('currency', 'CZK')


    if request.form.get('_method') == 'DELETE':
        delete_id = int(request.form['delete_id'])
        cursor = g.db.cursor()
        cursor.execute('SELECT ID FROM Prijmy WHERE ID = ? AND UserID = ?', (delete_id, current_user.id))
        income_data = cursor.fetchone()
        if income_data:
            cursor.execute('DELETE FROM Prijmy WHERE ID = ?', (delete_id,))
            g.db.commit()
            flash('Příjem byl úspěšně smazán.', 'success')
        else:
            flash('Nemáte oprávnění ke smazání tohoto příjmu.', 'error')
        return redirect(url_for('incomes'))

    if request.method == 'POST':
        if 'popis' in request.form:  # Kontrola, zda je pole "popis" ve formuláři k dispozici
            popis = request.form['popis']
            castka = float(request.form.get('částka', 0))
            datum = request.form['datum']

            if castka < 0:
                flash('Částka nemůže být záporná.', 'error')
                return redirect(url_for('incomes'))

            cursor = g.db.cursor()
            # Přepočet měny pro ukládání hodnoty v CZK
            if vybrana_mena != 'CZK':
                castka = prevest_menu(castka, vybrana_mena, to_czk=True)

            cursor.execute('INSERT INTO Prijmy (Popis, Castka, Datum, UserID) VALUES (?, ?, ?, ?)', (popis, castka, datum, current_user.id))
            g.db.commit()
            flash('Příjem byl úspěšně uložen.', 'success')
            return redirect(url_for('incomes'))
        else:
            new_currency = request.form['currency']
            session['currency'] = new_currency
            flash(f'Vybraná měna byla úspěšně změněna na {new_currency}.', 'success')
            return redirect(url_for('incomes'))

    cursor = g.db.cursor()
    cursor.execute('SELECT ID, Popis, Castka, Datum FROM Prijmy WHERE UserID = ?', (current_user.id,))
    incomes_data = cursor.fetchall()

    # Přepočet měn pro zobrazení v tabulce
    incomes_data_converted = [(income[0], income[1], prevest_menu(income[2], vybrana_mena), income[3]) for income in incomes_data]

    return render_template('incomes.html', incomes=incomes_data_converted, selected_currency=vybrana_mena, aktivni_str = aktivni_str)



@app.route('/expenses', methods=['GET', 'POST'])
@login_nutny
def expenses():
    aktivni_str = "expenses"

    vybrana_mena = session.get('currency', 'CZK')

    if request.form.get('_method') == 'DELETE':
        delete_id = int(request.form['delete_id'])
        cursor = g.db.cursor()

        cursor.execute('SELECT ID FROM Vydaje WHERE ID = ? AND UserID = ?', (delete_id, current_user.id))
        expense_data = cursor.fetchone()

        if expense_data:
            cursor.execute('DELETE FROM Vydaje WHERE ID = ?', (delete_id,))
            g.db.commit()
            flash('Výdaj byl úspěšně smazán.', 'success')
        else:
            flash('Nemáte oprávnění ke smazání tohoto výdaje.', 'error')

        return redirect(url_for('expenses'))

    if request.method == 'POST':
        if 'popis' in request.form:  # Kontrola, zda je pole "popis" ve formuláři k dispozici
            popis = request.form['popis']
            castka = float(request.form.get('částka', 0))
            datum = request.form['datum']
            kategorie = request.form['kategorie']

            if castka < 0:
                flash('Částka nemůže být záporná.', 'error')
                return redirect(url_for('expenses'))

            cursor = g.db.cursor()
            # Přepočet měny pro ukládání hodnoty v CZK
            if vybrana_mena != 'CZK':
                castka = prevest_menu(castka, vybrana_mena, to_czk=True)
                
            cursor.execute('INSERT INTO Vydaje (UserID, Kategorie, Popis, Castka, Datum) VALUES (?, ?, ?, ?, ?)',
                           (current_user.id, kategorie, popis, castka, datum))
            g.db.commit()
            flash('Výdaj byl úspěšně uložen.', 'success')
            return redirect(url_for('expenses'))
        else:
            new_currency = request.form['currency']
            session['currency'] = new_currency
            flash(f'Vybraná měna byla úspěšně změněna na {new_currency}.', 'success')
            return redirect(url_for('expenses'))

    cursor = g.db.cursor()

    # Zobrazujte pouze data přihlášeného uživatele
    cursor.execute('SELECT * FROM Vydaje WHERE UserID = ?', (current_user.id,))
    expenses_data = cursor.fetchall()

    expenses_data_converted = [(expense[0], expense[1], expense[2], prevest_menu(expense[3], vybrana_mena), expense[4]) for expense in expenses_data]

    return render_template('expenses.html', expenses=expenses_data_converted, selected_currency=vybrana_mena, aktivni_str = aktivni_str)


@app.route('/budget', methods=['GET', 'POST'])
@login_nutny
def budget():
    aktivni_str = "budget"
    vybrana_mena = session.get('currency', 'CZK')

    if request.form.get('_method') == 'DELETE':
        delete_id = int(request.form['delete_id'])
        cursor = g.db.cursor()

        cursor.execute('SELECT ID FROM Rozpocet WHERE ID = ? AND UserID = ?', (delete_id, current_user.id))
        expense_data = cursor.fetchone()

        if expense_data:
            cursor.execute('DELETE FROM Rozpocet WHERE ID = ?', (delete_id,))
            g.db.commit()
            flash('Rozpočet byl úspěšně smazán.', 'success')
        else:
            flash('Nemáte oprávnění ke smazání tohoto rozpočtu.', 'error')

        cursor.execute('DELETE FROM Rozpocet WHERE ID = ?', (delete_id,))
        g.db.commit()

        return redirect(url_for('budget'))

    if request.method == 'POST':
        if 'kategorie' in request.form:
            kategorie = request.form['kategorie']
            castka = float(request.form['částka'])
            mesic = request.form['měsíc']

            if castka < 0:
                flash('Částka nemůže být záporná.', 'error')
                return redirect(url_for('budget'))

            cursor = g.db.cursor()

            if vybrana_mena != 'CZK':
                castka = prevest_menu(castka, vybrana_mena, to_czk=True)

            cursor.execute('SELECT * FROM Rozpocet WHERE Kategorie = ? AND Mesic = ? AND UserID = ?', (kategorie, mesic, current_user.id))
            existing_budget = cursor.fetchone()

            if existing_budget:
                flash('Rozpočet pro zvolený měsíc a kategorii již existuje.', 'error')
                return redirect(url_for('budget'))

            cursor.execute('INSERT INTO Rozpocet (Kategorie, Castka, Mesic, UserID) VALUES (?, ?, ?, ?)',
                           (kategorie, castka, mesic, current_user.id))
            g.db.commit()
            flash('Rozpočet byl úspěšně uložen.', 'success')
            return redirect(url_for('budget'))
        else:
            new_currency = request.form['currency']
            session['currency'] = new_currency
            flash(f'Vybraná měna byla úspěšně změněna na {new_currency}.', 'success')
            return redirect(url_for('budget'))


    cursor = g.db.cursor()
    cursor.execute('SELECT * FROM Rozpocet WHERE UserID = ?', (current_user.id,))
    budget_data = cursor.fetchall()

    budget_data_converted = [(expense[0], expense[1], prevest_menu(expense[2], vybrana_mena), expense[3]) for expense in budget_data]

    return render_template('budget.html', budget_data=budget_data_converted, selected_currency=vybrana_mena, aktivni_str = aktivni_str)




@app.route('/budget_report/<int:budget_id>')
@login_nutny
def budget_report(budget_id):

    vybrana_mena = session.get('currency', 'CZK')
    cursor = g.db.cursor()

    cursor.execute('SELECT * FROM Rozpocet WHERE ID = ?', (budget_id,))
    budget_item = cursor.fetchone()

    if budget_item is None:
        return "Rozpočet neexistuje"

    cursor.execute('SELECT * FROM Vydaje WHERE Kategorie = ? AND UserID = ?', (budget_item[1], current_user.id,))
    expenses_for_budget = cursor.fetchall()

    budget_month = mesice.get(budget_item[3])

    related_expenses = []
    total_expense = 0

    for expense in expenses_for_budget:
        expense_date = datetime.strptime(expense[4], '%Y-%m-%d')
        expense_month = expense_date.month
        if int(expense_month) == int(budget_month):
            related_expenses.append(expense)
            total_expense += expense[3]

    total_budget = budget_item[2]
    remaining_budget = total_budget - total_expense

    # Výpočet procentuálního plnění rozpočtu
    used_percent = (total_expense / total_budget) * 100 if total_budget else 0

    report_html = "<div class='report-table'>"
    report_html += "<table>"
    report_html += "<tr><th>Popis</th><th>Částka</th><th>Datum</th></tr>"
    for expense in related_expenses:
        formatted_amount = "{:,.2f} {}".format(prevest_menu(expense[3], vybrana_mena), vybrana_mena).replace(",", " ")
        report_html += "<tr>"
        report_html += "<td style='max-width: 100px; word-wrap: break-word;'>{}</td><td>{}</td><td>{}</td>".format(expense[2], formatted_amount, expense[4])
        report_html += "</tr>"
    report_html += "</table>"
    report_html += "</div>"


    flex_data = [
        ("Výše rozpočtu:", prevest_menu(total_expense, vybrana_mena)),
        ("Celkové výdaje:", prevest_menu(total_budget, vybrana_mena)),
        ("Zbývá v rozpočtu:", prevest_menu(remaining_budget, vybrana_mena))
    ]

    report_html += "<div class='flex-container'>"
    for label, value in flex_data:
        value_formatted = "{:,.2f} {}".format(value, vybrana_mena).replace(",", " ")
        report_html += "<div class='cell'><p>{}</p><p>{}</p></div>".format(label, value_formatted)
    report_html += "</div>"

    # Přidání progress baru
    report_html += "<div class='budget-progress-bar'>"
    if remaining_budget > 0:
        report_html += "<div class='budget-progress' style='width: {}%'></div>".format(used_percent)
    else:
        report_html += "<div class='budget-progress-empty'></div>"
    report_html += "</div>"

    return report_html


# Funkce pro hashování hesla
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def send_email(to_email, subject, text_content):
    data = {
      'Messages': [
        {
          "From": {
            "Email": "andymonster57@gmail.com",
            "Name": "André Vágner"
          },
          "To": [
            {
              "Email": to_email
            }
          ],
          "Subject": subject,
          "TextPart": text_content,
        }
      ]
    }
    result = mailjet.send.create(data=data)
    return result.status_code

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['jmeno']
        password = request.form['heslo']
        email = request.form['email']

        # Hashování hesla před uložením do databáze
        hashed_password = hash_password(password)

        cursor = g.db.cursor()

        cursor.execute("SELECT * FROM Uzivatele WHERE Jmeno = ?", (username,))
        existing_username = cursor.fetchone()

        cursor.execute("SELECT * FROM Uzivatele WHERE Email = ?", (email,))
        existing_email = cursor.fetchone()

        if existing_username:
            flash('Uživatelské jméno již existuje!', 'error')
        elif existing_email:
            flash('E-mailová adresa je již obsazena!', 'error')

        else:

            subject = "Osobní finanční správce"
            text_content = "Děkujeme za registraci, {}! Nyní můžete začít aplikaci používat.".format(username)
            send_email(email, subject, text_content)

            cursor.execute("INSERT INTO Uzivatele (Jmeno, Heslo, Email) VALUES (?, ?, ?)", (username, hashed_password, email))
            g.db.commit()
            flash('Registrace byla úspěšná!', 'success')
            return redirect('/')

    return render_template('register.html')

# Model uživatele pro Flask-Login
class User(UserMixin):
    def __init__(self, id, jmeno):
        self.id = id
        self.jmeno = jmeno

# Funkce pro načtení uživatele podle ID (používáno Flask-Login)
@login_manager.user_loader
def load_user(user_id):
    cursor = g.db.cursor()
    cursor.execute("SELECT ID, Jmeno FROM Uzivatele WHERE ID = ?", (user_id,))
    user_data = cursor.fetchone()

    if user_data:
        user_id, user_name = user_data
        return User(user_id, user_name)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['jmeno']
        password = request.form['heslo']
        remember = 'rememberMe' in request.form

        if not (username and password):
            flash('Všechna pole jsou povinná!', 'error')
            return render_template('login.html')

        hashed_password = hash_password(password)

        cursor = g.db.cursor()
        cursor.execute("SELECT ID, Jmeno FROM Uzivatele WHERE Jmeno = ?", (username,))
        user_data = cursor.fetchone()

        if user_data:
            user_id, user_name = user_data

            cursor.execute("SELECT ID, Jmeno FROM Uzivatele WHERE Jmeno = ? AND Heslo = ?", (username, hashed_password))
            user_data_with_password = cursor.fetchone()

            if user_data_with_password:
                cursor.execute('UPDATE Uzivatele SET Mena = ? WHERE Jmeno = ? AND Mena IS NULL', ("CZK", username))
                cursor.execute('SELECT Mena FROM Uzivatele WHERE Jmeno = ?', (username,))
                mena_result = cursor.fetchone()

                if mena_result:
                    mena = mena_result[0]
                else:
                    mena = "CZK"

                session['currency'] = mena

                login_user(User(user_id, user_name), remember=remember)
                return redirect('/')
            else:
                flash('Chybné uživatelské jméno nebo heslo.', 'error')
        else:
            flash('Chybné uživatelské jméno nebo heslo.', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect('/login')


@app.route('/settings', methods=['GET', 'POST'])
@login_nutny
def settings():
    if request.method == 'POST':
        vybrana_mena = request.form.get('currency')
        if vybrana_mena:
            cursor = g.db.cursor()
            cursor.execute('UPDATE Uzivatele SET Mena = ? WHERE ID = ?', (vybrana_mena, current_user.id))
            g.db.commit()

            session['currency'] = vybrana_mena
            flash(f'Vybraná měna byla změněna na {vybrana_mena}.', 'info')

        return redirect(url_for('settings'))

    vybrana_mena = session.get('currency', 'CZK')
    return render_template('settings.html', selected_currency=vybrana_mena)


@app.route('/export_data', methods=['GET'])
def export_data():

    # Načtení aktuálně vybrané měny uživatele
    vybrana_mena = session.get('currency', 'CZK')

    # Vytvoření Excel workbook
    wb = Workbook()
    prijmy = wb.active
    prijmy.title = 'Prijmy'

    cursor = g.db.cursor()
    # Přidání hlaviček pro 'Prijmy'
    prijmy.append(['Popis', 'Částka', 'Datum'])
    cursor.execute('SELECT Popis, Castka, Datum FROM Prijmy WHERE UserID = ?', (current_user.id,))
    for row in cursor.fetchall():
        # Převedení částky do vybrané měny
        prevedena_castka = prevest_menu(row[1], vybrana_mena)
        prijmy.append([row[0], prevedena_castka, row[2]])

    # Přidání nového listu pro 'Vydaje' a zápis dat
    vydaje = wb.create_sheet(title='Vydaje')
    vydaje.append(['Kategorie', 'Popis', 'Částka', 'Datum'])
    cursor.execute('SELECT Kategorie, Popis, Castka, Datum FROM Vydaje WHERE UserID = ?', (current_user.id,))
    for row in cursor.fetchall():
        # Převedení částky do vybrané měny
        prevedena_castka = prevest_menu(row[2], vybrana_mena)
        vydaje.append([row[0], row[1], prevedena_castka, row[3]])

    # Přidání nového listu pro 'Rozpocty' a zápis dat
    rozpocty = wb.create_sheet(title='Rozpocty')
    rozpocty.append(['Kategorie', 'Částka', 'Měsíc'])
    cursor.execute('SELECT Kategorie, Castka, Mesic FROM Rozpocet WHERE UserID = ?', (current_user.id,))
    for row in cursor.fetchall():
        # Převedení částky do vybrané měny
        prevedena_castka = prevest_menu(row[1], vybrana_mena)
        rozpocty.append([row[0], prevedena_castka, row[2]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="all_data_{current_user.id}.xlsx"'}
    )

# Funkce pro změnu jména
@app.route('/change_name', methods=['POST'])
@login_nutny
def change_name():
    new_name = request.form['new_name']
    cursor = g.db.cursor()
    try:
        cursor.execute('UPDATE Uzivatele SET Jmeno = ? WHERE ID = ?', (new_name, current_user.id))
        g.db.commit()
        flash('Jméno bylo úspěšně změněno.','info')
    except Exception as e:
        flash(f'Chyba při změně jména: {e}','error')
    finally:
        cursor.close()
    return redirect(url_for('settings'))


# Funkce pro aktualizaci hesla
@app.route('/change_password', methods=['POST'])
@login_nutny
def change_password():
    old_password = request.form['old_password']
    new_password = request.form['new_password']
    hashed_old_password = hash_password(old_password)
    hashed_new_password = hash_password(new_password)

    cursor = g.db.cursor()

    cursor.execute('SELECT Heslo FROM Uzivatele WHERE ID = ?', (current_user.id,))
    current_password = cursor.fetchone()[0]

    if hashed_old_password != current_password:
        flash('Vaše staré heslo je nesprávné.', 'error')
        return redirect(url_for('settings'))

    try:
        cursor.execute('UPDATE Uzivatele SET Heslo = ? WHERE ID = ?', (hashed_new_password, current_user.id))
        g.db.commit()
        flash('Heslo bylo úspěšně změněno.', 'info')
    except Exception as e:
        flash(f'Chyba při změně hesla: {e}', 'error')
    finally:
        cursor.close()

    return redirect(url_for('settings'))

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
import numpy as np
from sklearn.linear_model import LinearRegression

@app.route('/get_data')
def get_data():
    cursor = g.db.cursor()
    vybrana_mena = session.get('currency', 'CZK')

    # Execute a query to retrieve the data
    cursor.execute('SELECT Datum, Castka FROM Prijmy WHERE UserID = ?', (current_user.id,))
    data_prijmy = [{'datum': row[0], 'castka': prevest_menu(row[1], vybrana_mena)} for row in cursor.fetchall()]

    cursor.execute('SELECT Datum, Castka FROM Vydaje WHERE UserID = ?', (current_user.id,))
    data_vydaje = [{'datum': row[0], 'castka': prevest_menu(row[1], vybrana_mena)} for row in cursor.fetchall()]

    # Třídění podle data
    data_prijmy.sort(key=lambda x: x['datum'])
    data_vydaje.sort(key=lambda x: x['datum'])

    # Suma po měsících
    summed_data_prijmy = defaultdict(float)
    for entry in data_prijmy:
        month = int(entry['datum'].split('-')[1])
        year = int(entry['datum'].split('-')[0])
        summed_data_prijmy[(year, month)] += entry['castka']

    summed_data_vydaje = defaultdict(float)
    for entry in data_vydaje:
        month = int(entry['datum'].split('-')[1])
        year = int(entry['datum'].split('-')[0])
        summed_data_vydaje[(year, month)] += entry['castka']

    existing_months_prijmy = set((year, month) for (year, month) in summed_data_prijmy.keys())
    existing_months_vydaje = set((year, month) for (year, month) in summed_data_vydaje.keys())

    # Seřazení podle roku a měsíce
    sorted_prijmy = sorted(existing_months_prijmy)
    sorted_vydaje = sorted(existing_months_vydaje)

    # Nejstarší a nejnovější datum
    min_date = min(sorted_prijmy[0], sorted_vydaje[0])
    max_date = max(sorted_prijmy[-1], sorted_vydaje[-1])

    all_dates_prijmy = set(sorted_prijmy)
    all_dates_vydaje = set(sorted_vydaje)

    # Doplnění chybějících dat mezi nejstarším a nejnovějším datem
    for year in range(min_date[0], max_date[0] + 1):
        for month in range(1, 13):
            current_date = (year, month)
            if min_date <= current_date <= max_date:
                if current_date not in all_dates_prijmy:
                    sorted_prijmy.append(current_date)
                if current_date not in all_dates_vydaje:
                    sorted_vydaje.append(current_date)

    real_prijmy_list = [
        {'datum': f"{month:02}/{year}", 'total': summed_data_prijmy.get((year, month), 0)}
        for year, month in sorted_prijmy
    ]
    real_vydaje_list = [
        {'datum': f"{month:02}/{year}", 'total': summed_data_vydaje.get((year, month), 0)}
        for year, month in sorted_vydaje
    ]

    real_prijmy_list = sorted(real_prijmy_list, key=lambda x: (int(x['datum'][-4:]), int(x['datum'][:2])))
    real_vydaje_list = sorted(real_vydaje_list, key=lambda x: (int(x['datum'][-4:]), int(x['datum'][:2])))

    rozdil_list = [{'datum': prijem['datum'], 'rozdil': prijem['total'] - vydaj['total']} for prijem, vydaj in zip(real_prijmy_list, real_vydaje_list)]

    # Lineární regrese pro příjmy
    x_prijmy = np.arange(len(real_prijmy_list)).reshape(-1, 1)
    y_prijmy = np.array([item['total'] for item in real_prijmy_list])
    reg_prijmy = LinearRegression().fit(x_prijmy, y_prijmy)

    # Lineární regrese pro výdaje
    x_vydaje = np.arange(len(real_vydaje_list)).reshape(-1, 1)
    y_vydaje = np.array([item['total'] for item in real_vydaje_list])
    reg_vydaje = LinearRegression().fit(x_vydaje, y_vydaje)
    last_date = real_prijmy_list[-1]['datum']
 
    # Příprava dat pro regresi pro následujících 6 měsíců
    future_months = [datetime.strptime(last_date, '%m/%Y') + timedelta(days=30 * i) for i in range(1, 7)]
    future_months_str = [date.strftime('%m/%Y') for date in future_months]

    # Generování seznamu regresních příjmů a výdajů
    regresni_prijmy_list = [{'datum': date.strftime('%m/%Y'), 'total': real_prijmy_list[-1]['total'] if i == len(real_prijmy_list) - 1 else None} for i, date in enumerate([datetime.strptime(item['datum'], '%m/%Y') for item in real_prijmy_list])]
    regresni_vydaje_list = [{'datum': date.strftime('%m/%Y'), 'total': real_vydaje_list[-1]['total'] if i == len(real_vydaje_list) - 1 else None} for i, date in enumerate([datetime.strptime(item['datum'], '%m/%Y') for item in real_vydaje_list])]
    print(regresni_prijmy_list)

    last_real_prijem = real_prijmy_list[-1]['total']
    last_real_vydaj = real_vydaje_list[-1]['total']
    last_real_rozdil = rozdil_list[-1]['rozdil']
    print(last_real_prijem)

    # Přidání regresních dat pro následujících 6 měsíců
    prijmy_adjust = last_real_prijem - round(reg_prijmy.predict([[len(real_prijmy_list) - 1]])[0], 2)
    vydaje_adjust = last_real_vydaj - round(reg_vydaje.predict([[len(real_vydaje_list) - 1]])[0], 2)
    rozdil_adjust = last_real_rozdil - round(reg_prijmy.predict([[len(real_prijmy_list) - 1]])[0] - reg_vydaje.predict([[len(real_vydaje_list) - 1]])[0], 2)
    print(prijmy_adjust)

    regresni_prijmy_list += [{'datum': date, 'total': round(reg_prijmy.predict([[i + len(real_prijmy_list)]])[0], 2) + prijmy_adjust} for i, date in enumerate(future_months_str)]
    regresni_vydaje_list += [{'datum': date, 'total': round(reg_vydaje.predict([[i + len(real_vydaje_list)]])[0], 2) + vydaje_adjust} for i, date in enumerate(future_months_str)]
    print(regresni_prijmy_list)
    pom = 0

    # Přesun negativní hodnoty do druhého listu
    for i in range(len(real_prijmy_list),len(regresni_prijmy_list)):
        if regresni_prijmy_list[i]['total'] and regresni_vydaje_list[i]['total']< 0:
            pom = regresni_prijmy_list[i]['total']
            regresni_prijmy_list[i]['total'] = abs(regresni_vydaje_list[i]['total'])
            regresni_vydaje_list[i]['total'] = abs(pom)
        else:
            if regresni_prijmy_list[i]['total'] < 0:
                regresni_vydaje_list[i]['total'] += abs(regresni_prijmy_list[i]['total'])
                regresni_prijmy_list[i]['total'] = 0

            elif regresni_vydaje_list[i]['total'] < 0:
                regresni_prijmy_list[i]['total'] += abs(regresni_prijmy_list[i]['total'])
                regresni_vydaje_list[i]['total'] = 0


    # Odečítání regrese příjmů a výdajů
    predpokladany_rozdil = [{'datum': date.strftime('%m/%Y'), 'total': last_real_rozdil if i == len(rozdil_list) - 1 else None} for i, date in enumerate([datetime.strptime(item['datum'], '%m/%Y') for item in real_prijmy_list])]
    # Odečítání regrese pro následujících 6 měsíců
    predpokladany_rozdil += [{'datum': date, 'total': round(reg_prijmy.predict([[i + len(real_prijmy_list)]])[0] - reg_vydaje.predict([[i + len(real_vydaje_list)]])[0], 2) + rozdil_adjust} for i, date in enumerate(future_months_str)]

    unique_dates = sorted(set(item['datum'] for item in real_prijmy_list + real_vydaje_list + predpokladany_rozdil),
                        key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])))

    return jsonify(real_prijmy=real_prijmy_list, real_vydaje=real_vydaje_list,
                regresni_prijmy=regresni_prijmy_list, regresni_vydaje=regresni_vydaje_list,
                unique_dates=unique_dates, celkove_prijmy=rozdil_list, predpokladany_rozdil=predpokladany_rozdil)


if __name__ == '__main__':
    app.run()